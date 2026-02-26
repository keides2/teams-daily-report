"""src/teams_chat_from_outlook_rerun.py

2月分など「特定期間だけ」Outlookの Teams日報フォルダーから再取り込みするための再実行用スクリプト。

- 元の `src/teams_chat_from_outlook.py` は変更しません
- 受信日(ReceivedTime)で期間フィルタします
- `--reprocess` を指定すると、processed_mail_ids.json に載っているメールでも再処理します

使い方:
  python src\\teams_chat_from_outlook_rerun.py --since 2026-02-01 --until 2026-02-28 --reprocess

注意:
- 既にExcelに書き込まれている内容がある場合、重複して追記される可能性があります。
  必要に応じて実行前にExcel(2月シート)の該当セルを確認してください。
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
import win32com.client


def parse_yyyy_mm_dd(s: str) -> dt.date:
    try:
        return dt.date.fromisoformat(s)
    except ValueError as e:
        raise argparse.ArgumentTypeError(
            "日付は YYYY-MM-DD 形式で指定してください (例: 2026-02-01)"
        ) from e


def load_state(state_file: Path) -> set[str]:
    if state_file.exists():
        return set(json.loads(state_file.read_text(encoding="utf-8")))
    return set()


def save_state(state_file: Path, ids: set[str]) -> None:
    state_file.write_text(
        json.dumps(list(ids), ensure_ascii=False, indent=2), encoding="utf-8"
    )


def extract_daily_reports(text: str, mail_received_date: dt.date | None = None):
    """return list of (kind, summary, date) tuples."""
    if "Microsoft Teams" in text:
        parts = text.split("Microsoft Teams")
        if len(parts) > 1:
            text = parts[-1]

    reports: list[tuple[str, str, dt.date | None]] = []

    base_year = mail_received_date.year if mail_received_date else dt.date.today().year

    for m in re.finditer(
        r"#日報計画\s+(\d{1,2})/(\d{1,2})\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)",
        text,
    ):
        month = int(m.group(1))
        day = int(m.group(2))
        summary = m.group(3).strip()[:50]
        if not summary:
            continue
        try:
            year = base_year
            if mail_received_date:
                if mail_received_date.month == 12 and month == 1:
                    year = base_year + 1
                elif mail_received_date.month == 1 and month == 12:
                    year = base_year - 1
            report_date = dt.date(year, month, day)
            reports.append(("plan", summary, report_date))
        except ValueError:
            print(f"  警告: 無効な日付 {month}/{day}")

    for m in re.finditer(
        r"#日報結果\s+(\d{1,2})/(\d{1,2})\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)",
        text,
    ):
        month = int(m.group(1))
        day = int(m.group(2))
        summary = m.group(3).strip()[:50]
        if not summary:
            continue
        try:
            year = base_year
            if mail_received_date:
                if mail_received_date.month == 12 and month == 1:
                    year = base_year + 1
                elif mail_received_date.month == 1 and month == 12:
                    year = base_year - 1
            report_date = dt.date(year, month, day)
            reports.append(("result", summary, report_date))
        except ValueError:
            print(f"  警告: 無効な日付 {month}/{day}")

    for m in re.finditer(r"#日報\s+(\d{1,2})/(\d{1,2})\s+(.+?)(?:\n|$)", text):
        month = int(m.group(1))
        day = int(m.group(2))
        summary = m.group(3).strip()[:50]
        if summary and "計画" not in summary and "結果" not in summary:
            try:
                year = base_year
                if mail_received_date:
                    if mail_received_date.month == 12 and month == 1:
                        year = base_year + 1
                    elif mail_received_date.month == 1 and month == 12:
                        year = base_year - 1
                report_date = dt.date(year, month, day)
                reports.append(("result", summary, report_date))
            except ValueError:
                print(f"  警告: 無効な日付 {month}/{day}")

    if not reports:
        for m in re.finditer(r"#日報計画\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
            summary = m.group(1).strip()[:50]
            if summary and not re.match(r"\d{1,2}/\d{1,2}", summary):
                reports.append(("plan", summary, None))

        for m in re.finditer(r"#日報結果\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
            summary = m.group(1).strip()[:50]
            if summary and not re.match(r"\d{1,2}/\d{1,2}", summary):
                reports.append(("result", summary, None))

        for m in re.finditer(r"#日報\s+(.+?)(?:\n|$)", text):
            summary = m.group(1).strip()[:50]
            if (
                summary
                and "計画" not in summary
                and "結果" not in summary
                and not re.match(r"\d{1,2}/\d{1,2}", summary)
            ):
                reports.append(("result", summary, None))

    return reports


def find_row_by_date(ws, date_col: str, target_date: dt.date):
    for r in range(1, ws.max_row + 1):
        v = ws[f"{date_col}{r}"].value
        if isinstance(v, (dt.date, dt.datetime)):
            cell_date = v.date() if isinstance(v, dt.datetime) else v
            if cell_date == target_date:
                return r
    return None


def resolve_excel_path(today: dt.date) -> Path:
    year = today.year
    month = today.month
    fiscal_year = year if month >= 4 else year - 1

    if month >= 4:
        fiscal_month_num = month - 3
    else:
        fiscal_month_num = month + 9

    month_folder = f"{fiscal_month_num:02d}_{month}月"
    month_name = f"{month}月"

    excel_path_template = os.getenv("EXCEL_PATH", "")
    if excel_path_template:
        excel_path_str = excel_path_template.format(
            fiscal_year=fiscal_year, month_folder=month_folder, month_name=month_name
        )
        return Path(excel_path_str)

    username = os.getenv("EXCEL_USERNAME") or os.getenv("USERNAME", "username")
    company = os.getenv("EXCEL_COMPANY", "株式会社サンプル")
    department = os.getenv("EXCEL_DEPARTMENT", "部署名-01.庶務事項")
    category = os.getenv("EXCEL_CATEGORY", "業務管理")
    base_folder = os.getenv("EXCEL_BASE_FOLDER", "業務内容報告書")
    team = os.getenv("EXCEL_TEAM", "チームA")
    user_name = os.getenv("EXCEL_USER_NAME", "山田太郎")

    onedrive_base = f"C:\\Users\\{username}\\OneDrive - {company}"

    return Path(
        f"{onedrive_base}\\{department}\\{category}\\"
        f"{base_folder}\\{fiscal_year}年度\\{month_folder}\\{team}\\"
        f"（{month_name}_{user_name}）業務内容報告書.xlsm"
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="OutlookのTeams日報フォルダーから特定期間だけ再取り込みしてExcelへ書き込みます"
    )
    parser.add_argument("--since", type=parse_yyyy_mm_dd, required=True)
    parser.add_argument("--until", type=parse_yyyy_mm_dd, required=True)
    parser.add_argument("--reprocess", action="store_true")
    parser.add_argument(
        "--debug-list",
        type=int,
        default=20,
        help="件数差異の切り分け用に、先頭N件の受信日/件名を表示します(0で無効)",
    )
    args = parser.parse_args()

    load_dotenv()

    outlook_folder = os.getenv("OUTLOOK_FOLDER", "Teams日報")
    state_file = Path("processed_mail_ids.json")

    date_col = os.getenv("DATE_COL", "B")
    plan_col = os.getenv("PLAN_COL", "C")
    result_col = os.getenv("RESULT_COL", "F")
    rows_per_day = int(os.getenv("ROWS_PER_DAY", "6"))

    today = dt.date.today()
    excel_path = resolve_excel_path(today)

    print(f"対象フォルダー: {outlook_folder}")
    print(f"対象期間    : {args.since} 〜 {args.until} (受信日 기준)")
    print(f"再処理      : {args.reprocess}")
    print(f"Excel       : {excel_path}")
    print()

    if not excel_path.exists():
        raise RuntimeError(f"Excelファイルが見つかりません: {excel_path}")

    processed = load_state(state_file)
    new_processed = set(processed)

    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
    inbox = namespace.GetDefaultFolder(6)

    target_folder = None
    for f in inbox.Folders:
        if f.Name == outlook_folder:
            target_folder = f
            break
    if not target_folder:
        raise RuntimeError(f"Outlook フォルダーが見つかりません: {outlook_folder}")

    # Outlook 側の診断情報（Store/フォルダパスが想定通りか確認する）
    try:
        store_name = target_folder.Store.DisplayName
    except Exception:
        store_name = "(unknown)"
    try:
        folder_path = target_folder.FolderPath
    except Exception:
        folder_path = "(unknown)"

    print(f"Outlook Store : {store_name}")
    print(f"FolderPath   : {folder_path}")

    mail_items = target_folder.Items
    # Items は並びが不定で取りこぼしが出ることがあるため、Sortして安定化してから走査する
    try:
        mail_items.Sort("[ReceivedTime]", True)  # True=降順
    except Exception:
        pass

    print(f"Items.Count  : {mail_items.Count}")
    print()

    try:
        wb = load_workbook(excel_path, keep_vba=True, data_only=False)
    except PermissionError:
        raise RuntimeError("Excelファイルが開かれています。閉じてから再実行してください。")

    # 当月シート選択（元スクリプトと同じ「◯月」を含むシート優先）
    sheet_name_pattern = f"{today.month}月"
    ws = wb.active
    for sheet_name in wb.sheetnames:
        if sheet_name_pattern in sheet_name:
            ws = wb[sheet_name]
            break

    processed_count = 0
    new_count = 0
    in_range_count = 0
    enumerated_count = 0

    debug_rows: list[tuple[dt.date | None, str]] = []

    for mail in mail_items:
        enumerated_count += 1
        entry_id = mail.EntryID

        received = mail.ReceivedTime
        received_date: dt.date | None
        if isinstance(received, dt.datetime):
            received_date = received.date()
        else:
            received_date = None

        if args.debug_list and len(debug_rows) < args.debug_list:
            subject = (mail.Subject or "(件名なし)")
            debug_rows.append((received_date, subject))

        if received_date is None:
            continue

        if received_date < args.since or received_date > args.until:
            continue
        in_range_count += 1

        if entry_id in processed and not args.reprocess:
            processed_count += 1
            continue

        body = mail.Body or ""
        subject = mail.Subject or "(件名なし)"

        reports = extract_daily_reports(body, received_date)
        if not reports:
            continue

        print(f"✓ 処理: {subject[:50]} ({received_date})")

        # 日付ごとにグルーピング
        date_groups: dict[dt.date, list[tuple[str, str]]] = {}
        for kind, summary, report_date in reports:
            target_date = report_date if report_date else received_date
            date_groups.setdefault(target_date, []).append((kind, summary))

        for target_date, day_items in date_groups.items():
            date_row = find_row_by_date(ws, date_col, target_date)
            if date_row is None:
                print(f"  ✗ スキップ: {target_date} の日付行が見つかりません")
                continue

            input_start_row = date_row - 3

            # その日付ブロック内で次の空き行を追跡
            plan_row_offset = 0
            result_row_offset = 0

            for kind, summary in day_items:
                if kind == "plan":
                    written = False
                    for offset in range(plan_row_offset, rows_per_day):
                        r = input_start_row + offset
                        cell = ws[f"{plan_col}{r}"]
                        if not cell.value:
                            cell.value = summary
                            written = True
                            plan_row_offset = offset + 1
                            print(f"  → {target_date} plan: {plan_col}{r}")
                            break
                    if not written:
                        r = input_start_row + rows_per_day - 1
                        cell = ws[f"{plan_col}{r}"]
                        cell.value = f"{cell.value}\n{summary}" if cell.value else summary
                        print(f"  → {target_date} plan(追記): {plan_col}{r}")

                else:
                    written = False
                    for offset in range(result_row_offset, rows_per_day):
                        r = input_start_row + offset
                        cell = ws[f"{result_col}{r}"]
                        if not cell.value:
                            cell.value = summary
                            written = True
                            result_row_offset = offset + 1
                            print(f"  → {target_date} result: {result_col}{r}")
                            break
                    if not written:
                        r = input_start_row + rows_per_day - 1
                        cell = ws[f"{result_col}{r}"]
                        cell.value = f"{cell.value}\n{summary}" if cell.value else summary
                        print(f"  → {target_date} result(追記): {result_col}{r}")

    new_processed.add(entry_id)
    new_count += 1

    print()
    print("--- Debug (先頭N件) ---")
    if args.debug_list:
        for d, subject in debug_rows:
            ds = d.isoformat() if d else "(no date)"
            print(f"{ds} | {subject[:80]}")
        print()

    print("--- 集計 ---")
    print(f"Items.Count        : {mail_items.Count}")
    print(f"列挙できた件数     : {enumerated_count}")
    print(f"期間内メール数     : {in_range_count}")
    print(f"処理済みスキップ   : {processed_count}")
    print(f"今回処理           : {new_count}")

    wb.save(excel_path)
    save_state(state_file, new_processed)
    print("\n日報更新完了")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
