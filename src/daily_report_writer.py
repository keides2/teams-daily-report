"""
CLI 直接書き込みモード
  Outlook メール取得をスキップし、コマンドライン引数から直接 Excel 日報に書き込む。
  Copilot Chat スラッシュコマンド (/daily-plan, /daily-result) から呼び出される。

使用例:
  python src/daily_report_writer.py --mode plan --summary "設計レビューの準備"
  python src/daily_report_writer.py --mode result --date 2/21 --summary "テスト完了"

引数:
  --mode    plan=計画列(C), result=実績列(F)
  --summary Excel に書き込む文字列（50文字以内に自動切り詰め）
  --date    書き込む日付 MM/DD 形式（省略時は今日）
"""

import argparse
import json
import os
import re
import datetime as dt
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

# .env ファイルから環境変数を読み込み
load_dotenv()

# ===== 設定 =====
OUTLOOK_FOLDER = os.getenv("OUTLOOK_FOLDER", "Teams日報")
STATE_FILE = Path("processed_mail_ids.json")

# 現在の月に基づいてファイルパスを動的に生成
today = dt.date.today()
year = today.year
month = today.month

# 年度計算（4月始まり）
fiscal_year = year if month >= 4 else year - 1

# 年度内の月番号（4月=01, 5月=02, ..., 12月=09, 1月=10, 2月=11, 3月=12）
if month >= 4:
    fiscal_month_num = month - 3  # 4月=1, 5月=2, ..., 12月=9
else:
    fiscal_month_num = month + 9  # 1月=10, 2月=11, 3月=12

# 月フォルダ名（09_12月 形式）
month_folder = f"{fiscal_month_num:02d}_{month}月"

# ファイル名の月部分
month_name = f"{month}月"

# ===== Excelファイルパス設定 =====
# 環境変数 EXCEL_PATH が設定されていればそれを使用
excel_path_template = os.getenv("EXCEL_PATH", "")

if excel_path_template:
    # テンプレート文字列の置換
    excel_path_str = excel_path_template.format(
        fiscal_year=fiscal_year,
        month_folder=month_folder,
        month_name=month_name
    )
    EXCEL_PATH = Path(excel_path_str)
else:
    # 環境変数からテンプレート要素を取得してパスを生成
    username = os.getenv("EXCEL_USERNAME") or os.getenv("USERNAME", "username")
    company = os.getenv("EXCEL_COMPANY", "株式会社サンプル")
    department = os.getenv("EXCEL_DEPARTMENT", "部署名-01.庶務事項")
    category = os.getenv("EXCEL_CATEGORY", "業務管理")
    base_folder = os.getenv("EXCEL_BASE_FOLDER", "業務内容報告書")
    team = os.getenv("EXCEL_TEAM", "チームA")
    user_name = os.getenv("EXCEL_USER_NAME", "山田太郎")
    
    # OneDriveベースパス
    onedrive_base = f"C:\\Users\\{username}\\OneDrive - {company}"
    
    EXCEL_PATH = Path(
        f"{onedrive_base}\\{department}\\{category}\\"
        f"{base_folder}\\{fiscal_year}年度\\{month_folder}\\{team}\\"
        f"（{month_name}_{user_name}）業務内容報告書.xlsm"
    )

# Excel構造設定
SHEET_NAME_PATTERN = f"{month}月"
DATE_COL = os.getenv("DATE_COL", "B")
PLAN_COL = os.getenv("PLAN_COL", "C")
RESULT_COL = os.getenv("RESULT_COL", "F")
ROWS_PER_DAY = int(os.getenv("ROWS_PER_DAY", "6"))


# ===== ユーティリティ =====
def load_state():
    if STATE_FILE.exists():
        return set(json.loads(STATE_FILE.read_text(encoding="utf-8")))
    return set()

def save_state(ids):
    STATE_FILE.write_text(json.dumps(list(ids), ensure_ascii=False, indent=2), encoding="utf-8")

def extract_daily_reports(text: str, mail_received_date=None):
    """
    return list of (kind, summary, date) tuples
    1つのメッセージから複数の日報（計画・結果）を抽出
    
    対応パターン:
    1. #日報計画 MM/DD 要約:xxxxx
    2. #日報結果 MM/DD 要約:xxxxx
    3. #日報計画 MM/DD xxxxx
    4. #日報結果 MM/DD xxxxx
    5. #日報 MM/DD xxxxx (結果として扱う)
    6. 日付なしの場合は None を返す（メール受信日を使用）
    
    Note: Teamsのメール本文にはプレビュー部分と実際のメッセージ部分が含まれる。
          実際のメッセージ部分のみを抽出するため、"Microsoft Teams" 以降を使用。
    """
    # Teamsメールの実際のメッセージ部分を抽出
    # "Microsoft Teams" というテキストの後が実際のメッセージ
    if "Microsoft Teams" in text:
        # "Microsoft Teams" の後、発信者名の後から実際のメッセージが始まる
        parts = text.split("Microsoft Teams")
        if len(parts) > 1:
            # 後半部分のみを使用（実際のチャットメッセージ）
            text = parts[-1]
    
    reports = []
    # メール受信日がある場合はその年を基準に、ない場合は今日の年を使用
    if mail_received_date:
        base_year = mail_received_date.year
    else:
        base_year = dt.date.today().year
    
    # #日報計画を全て抽出（日付付き）
    for m in re.finditer(r"#日報計画\s+(\d{1,2})/(\d{1,2})\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
        month = int(m.group(1))
        day = int(m.group(2))
        summary = m.group(3).strip()[:50]
        if summary:
            try:
                # 年をまたぐ場合の処理（メール受信日を基準に判定）
                if mail_received_date:
                    # メール受信月と報告月を比較して年を決定
                    if mail_received_date.month == 12 and month == 1:
                        # 12月に受信して1月の日付を指定した場合は翌年
                        year = base_year + 1
                    elif mail_received_date.month == 1 and month == 12:
                        # 1月に受信して12月の日付を指定した場合は前年
                        year = base_year - 1
                    else:
                        year = base_year
                else:
                    year = base_year
                report_date = dt.date(year, month, day)
                reports.append(("plan", summary, report_date))
            except ValueError:
                print(f"  警告: 無効な日付 {month}/{day}")
    
    # #日報結果を全て抽出（日付付き）
    for m in re.finditer(r"#日報結果\s+(\d{1,2})/(\d{1,2})\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
        month = int(m.group(1))
        day = int(m.group(2))
        summary = m.group(3).strip()[:50]
        if summary:
            try:
                # 年をまたぐ場合の処理（メール受信日を基準に判定）
                if mail_received_date:
                    # メール受信月と報告月を比較して年を決定
                    if mail_received_date.month == 12 and month == 1:
                        # 12月に受信して1月の日付を指定した場合は翌年
                        year = base_year + 1
                    elif mail_received_date.month == 1 and month == 12:
                        # 1月に受信して12月の日付を指定した場合は前年
                        year = base_year - 1
                    else:
                        year = base_year
                else:
                    year = base_year
                report_date = dt.date(year, month, day)
                reports.append(("result", summary, report_date))
            except ValueError:
                print(f"  警告: 無効な日付 {month}/{day}")
    
    # #日報（単独）を全て抽出（日付付き、結果として扱う）
    for m in re.finditer(r"#日報\s+(\d{1,2})/(\d{1,2})\s+(.+?)(?:\n|$)", text):
        month = int(m.group(1))
        day = int(m.group(2))
        summary = m.group(3).strip()[:50]
        if summary and "計画" not in summary and "結果" not in summary:
            try:
                # 年をまたぐ場合の処理（メール受信日を基準に判定）
                if mail_received_date:
                    # メール受信月と報告月を比較して年を決定
                    if mail_received_date.month == 12 and month == 1:
                        # 12月に受信して1月の日付を指定した場合は翌年
                        year = base_year + 1
                    elif mail_received_date.month == 1 and month == 12:
                        # 1月に受信して12月の日付を指定した場合は前年
                        year = base_year - 1
                    else:
                        year = base_year
                else:
                    year = base_year
                report_date = dt.date(year, month, day)
                reports.append(("result", summary, report_date))
            except ValueError:
                print(f"  警告: 無効な日付 {month}/{day}")
    
    # 日付なしのパターン（後方互換性のため）
    if not reports:
        # #日報計画（日付なし）
        for m in re.finditer(r"#日報計画\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
            summary = m.group(1).strip()[:50]
            if summary and not re.match(r"\d{1,2}/\d{1,2}", summary):
                reports.append(("plan", summary, None))
        
        # #日報結果（日付なし）
        for m in re.finditer(r"#日報結果\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
            summary = m.group(1).strip()[:50]
            if summary and not re.match(r"\d{1,2}/\d{1,2}", summary):
                reports.append(("result", summary, None))
        
        # #日報（日付なし）
        for m in re.finditer(r"#日報\s+(.+?)(?:\n|$)", text):
            summary = m.group(1).strip()[:50]
            if (summary and "計画" not in summary and
                    "結果" not in summary and
                    not re.match(r"\d{1,2}/\d{1,2}", summary)):
                reports.append(("result", summary, None))
    
    return reports


def find_row_by_date(ws, target_date):
    """B列から target_date に一致する行番号を返す。見つからなければ None。"""
    for r in range(1, ws.max_row + 1):
        v = ws[f"{DATE_COL}{r}"].value
        if isinstance(v, (dt.date, dt.datetime)):
            cell_date = v.date() if isinstance(v, dt.datetime) else v
            if cell_date == target_date:
                return r
    return None


def write_to_excel(target_date: dt.date, mode: str, summary: str) -> bool:
    """
    Excel の計画列(C) または実績列(F) に summary を書き込む。
    mode: 'plan' → PLAN_COL, 'result' → RESULT_COL
    """
    if not EXCEL_PATH.exists():
        print(f"エラー: Excelファイルが見つかりません: {EXCEL_PATH}")
        return False

    keep_vba = EXCEL_PATH.suffix.lower() == ".xlsm"
    try:
        wb = load_workbook(EXCEL_PATH, keep_vba=keep_vba, data_only=False)
    except PermissionError:
        print("エラー: Excelファイルが開かれています。閉じてから再実行してください。")
        return False
    except Exception as e:
        print(f"エラー: Excelファイルの読み込みに失敗しました。{type(e).__name__}: {e}")
        return False

    # シート選択
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if SHEET_NAME_PATTERN in sheet_name:
            target_sheet = sheet_name
            break
    if target_sheet:
        ws = wb[target_sheet]
    else:
        ws = wb.active
        print(f"警告: '{SHEET_NAME_PATTERN}'を含むシートが見つかりません。'{ws.title}'を使用します。")

    # 日付行を検索
    date_row = find_row_by_date(ws, target_date)
    if date_row is None:
        print(f"エラー: {target_date} ({target_date.month}/{target_date.day}) の日付行が見つかりません。")
        return False

    # 入力開始行（日付行の 3 行上）
    input_start_row = date_row - 3
    col = PLAN_COL if mode == "plan" else RESULT_COL

    # result モードは先頭行が計画列の数式コピーで埋まるため 2 行目から開始
    start_offset = 1 if mode == "result" else 0

    # 空き行を探して書き込み
    written = False
    for offset in range(start_offset, ROWS_PER_DAY):
        target_row = input_start_row + offset
        cell = ws[f"{col}{target_row}"]
        if not cell.value:
            cell.value = summary
            print(f"書き込み: {col}{target_row} ← {summary}")
            written = True
            break

    if not written:
        # 空き行がない場合は最終行に追記
        target_row = input_start_row + ROWS_PER_DAY - 1
        cell = ws[f"{col}{target_row}"]
        cell.value = f"{cell.value}\n{summary}" if cell.value else summary
        print(f"追記（空き行なし）: {col}{target_row} ← {summary}")

    # 保存
    try:
        wb.save(EXCEL_PATH)
        return True
    except PermissionError:
        print("エラー: Excelファイルの保存に失敗しました。ファイルを閉じてから再実行してください。")
        return False
    except Exception as e:
        print(f"エラー: Excelファイルの保存に失敗しました。{type(e).__name__}: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(
        description="Excel 日報に直接書き込む CLI ツール",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "例:\n"
            "  python src/daily_report_writer.py --mode plan --summary \"設計レビューの準備\"\n"
            "  python src/daily_report_writer.py --mode result --date 2/21 --summary \"テスト完了\""
        ),
    )
    parser.add_argument(
        "--mode", choices=["plan", "result"], required=True,
        help="plan=計画列(C), result=実績列(F)"
    )
    parser.add_argument(
        "--summary", required=True,
        help="Excel に書き込む文字列（50文字以内に自動切り詰め）"
    )
    parser.add_argument(
        "--date",
        help="書き込む日付 MM/DD 形式（省略時は今日）"
    )
    args = parser.parse_args()

    summary = args.summary.strip()[:50]

    # 日付解析
    if args.date:
        try:
            parts = args.date.split("/")
            m, d = int(parts[0]), int(parts[1])
            target_date = dt.date(today.year, m, d)
        except (ValueError, IndexError):
            print("エラー: 日付の形式が正しくありません。MM/DD 形式で指定してください。例: 2/21")
            exit(1)
    else:
        target_date = today

    mode_label = "計画 (C列)" if args.mode == "plan" else "実績 (F列)"
    print(f"モード  : {mode_label}")
    print(f"日付    : {target_date}")
    print(f"要約    : {summary}")
    print(f"Excel   : {EXCEL_PATH}")
    print()

    ok = write_to_excel(target_date, args.mode, summary)
    if ok:
        print("\n日報更新完了")
    else:
        exit(1)


if __name__ == "__main__":
    main()
