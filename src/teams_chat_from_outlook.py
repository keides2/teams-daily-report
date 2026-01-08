"""
計画
1) Outlook ローカルから Teams共有メールを取得
2) #日報計画 / #日報結果 + 要約: を抽出
3) 未処理メールのみ処理（EntryID で重複防止）
4) Excel 日報に追記
"""

import json
import os
import re
import datetime as dt
from pathlib import Path

from dotenv import load_dotenv
import win32com.client
from openpyxl import load_workbook

# .env ファイルから環境変数を読み込み
load_dotenv()

# ===== 設定 =====
OUTLOOK_FOLDER = os.getenv("OUTLOOK_FOLDER", "Teams日報")
STATE_FILE = Path("processed_mail_ids.json")

# 現在の月に基づいてファイルパスを動的に生成
today = dt.date.today()
year = today.year
month = today.month

# 年度計算（4月始まり）
fiscal_year = year if month >= 4 else year - 1

# 年度内の月番号（4月=01, 5月=02, ..., 12月=09, 1月=10, 2月=11, 3月=12）
if month >= 4:
    fiscal_month_num = month - 3  # 4月=1, 5月=2, ..., 12月=9
else:
    fiscal_month_num = month + 9  # 1月=10, 2月=11, 3月=12

# 月フォルダ名（09_12月 形式）
month_folder = f"{fiscal_month_num:02d}_{month}月"

# ファイル名の月部分
month_name = f"{month}月"

# ===== Excelファイルパス設定 =====
# 環境変数 EXCEL_PATH が設定されていればそれを使用
excel_path_template = os.getenv("EXCEL_PATH", "")

if excel_path_template:
    # テンプレート文字列の置換
    excel_path_str = excel_path_template.format(
        fiscal_year=fiscal_year,
        month_folder=month_folder,
        month_name=month_name
    )
    EXCEL_PATH = Path(excel_path_str)
else:
    # 環境変数からテンプレート要素を取得してパスを生成
    username = os.getenv("EXCEL_USERNAME") or os.getenv("USERNAME", "username")
    company = os.getenv("EXCEL_COMPANY", "株式会社サンプル")
    department = os.getenv("EXCEL_DEPARTMENT", "部署名-01.庶務事項")
    category = os.getenv("EXCEL_CATEGORY", "業務管理")
    base_folder = os.getenv("EXCEL_BASE_FOLDER", "業務内容報告書")
    team = os.getenv("EXCEL_TEAM", "チームA")
    user_name = os.getenv("EXCEL_USER_NAME", "山田太郎")
    
    # OneDriveベースパス
    onedrive_base = f"C:\\Users\\{username}\\OneDrive - {company}"
    
    EXCEL_PATH = Path(
        f"{onedrive_base}\\{department}\\{category}\\"
        f"{base_folder}\\{fiscal_year}年度\\{month_folder}\\{team}\\"
        f"（{month_name}_{user_name}）業務内容報告書.xlsm"
    )

# Excel構造設定
SHEET_NAME_PATTERN = f"{month}月"
DATE_COL = os.getenv("DATE_COL", "B")
PLAN_COL = os.getenv("PLAN_COL", "C")
RESULT_COL = os.getenv("RESULT_COL", "F")
ROWS_PER_DAY = int(os.getenv("ROWS_PER_DAY", "6"))


# ===== ユーティリティ =====
def load_state():
    if STATE_FILE.exists():
        return set(json.loads(STATE_FILE.read_text(encoding="utf-8")))
    return set()

def save_state(ids):
    STATE_FILE.write_text(json.dumps(list(ids), ensure_ascii=False, indent=2), encoding="utf-8")

def extract_daily_reports(text: str, mail_received_date=None):
    """
    return list of (kind, summary, date) tuples
    1つのメッセージから複数の日報（計画・結果）を抽出
    
    対応パターン:
    1. #日報計画 MM/DD 要約:xxxxx
    2. #日報結果 MM/DD 要約:xxxxx
    3. #日報計画 MM/DD xxxxx
    4. #日報結果 MM/DD xxxxx
    5. #日報 MM/DD xxxxx (結果として扱う)
    6. 日付なしの場合は None を返す（メール受信日を使用）
    
    Note: Teamsのメール本文にはプレビュー部分と実際のメッセージ部分が含まれる。
          実際のメッセージ部分のみを抽出するため、"Microsoft Teams" 以降を使用。
    """
    # Teamsメールの実際のメッセージ部分を抽出
    # "Microsoft Teams" というテキストの後が実際のメッセージ
    if "Microsoft Teams" in text:
        # "Microsoft Teams" の後、発信者名の後から実際のメッセージが始まる
        parts = text.split("Microsoft Teams")
        if len(parts) > 1:
            # 後半部分のみを使用（実際のチャットメッセージ）
            text = parts[-1]
    
    reports = []
    # メール受信日がある場合はその年を基準に、ない場合は今日の年を使用
    if mail_received_date:
        base_year = mail_received_date.year
    else:
        base_year = dt.date.today().year
    
    # #日報計画を全て抽出（日付付き）
    for m in re.finditer(r"#日報計画\s+(\d{1,2})/(\d{1,2})\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
        month = int(m.group(1))
        day = int(m.group(2))
        summary = m.group(3).strip()[:50]
        if summary:
            try:
                # 年をまたぐ場合の処理（メール受信日を基準に判定）
                if mail_received_date:
                    # メール受信月と報告月を比較して年を決定
                    if mail_received_date.month == 12 and month == 1:
                        # 12月に受信して1月の日付を指定した場合は翌年
                        year = base_year + 1
                    elif mail_received_date.month == 1 and month == 12:
                        # 1月に受信して12月の日付を指定した場合は前年
                        year = base_year - 1
                    else:
                        year = base_year
                else:
                    year = base_year
                report_date = dt.date(year, month, day)
                reports.append(("plan", summary, report_date))
            except ValueError:
                print(f"  警告: 無効な日付 {month}/{day}")
    
    # #日報結果を全て抽出（日付付き）
    for m in re.finditer(r"#日報結果\s+(\d{1,2})/(\d{1,2})\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
        month = int(m.group(1))
        day = int(m.group(2))
        summary = m.group(3).strip()[:50]
        if summary:
            try:
                # 年をまたぐ場合の処理（メール受信日を基準に判定）
                if mail_received_date:
                    # メール受信月と報告月を比較して年を決定
                    if mail_received_date.month == 12 and month == 1:
                        # 12月に受信して1月の日付を指定した場合は翌年
                        year = base_year + 1
                    elif mail_received_date.month == 1 and month == 12:
                        # 1月に受信して12月の日付を指定した場合は前年
                        year = base_year - 1
                    else:
                        year = base_year
                else:
                    year = base_year
                report_date = dt.date(year, month, day)
                reports.append(("result", summary, report_date))
            except ValueError:
                print(f"  警告: 無効な日付 {month}/{day}")
    
    # #日報（単独）を全て抽出（日付付き、結果として扱う）
    for m in re.finditer(r"#日報\s+(\d{1,2})/(\d{1,2})\s+(.+?)(?:\n|$)", text):
        month = int(m.group(1))
        day = int(m.group(2))
        summary = m.group(3).strip()[:50]
        if summary and "計画" not in summary and "結果" not in summary:
            try:
                # 年をまたぐ場合の処理（メール受信日を基準に判定）
                if mail_received_date:
                    # メール受信月と報告月を比較して年を決定
                    if mail_received_date.month == 12 and month == 1:
                        # 12月に受信して1月の日付を指定した場合は翌年
                        year = base_year + 1
                    elif mail_received_date.month == 1 and month == 12:
                        # 1月に受信して12月の日付を指定した場合は前年
                        year = base_year - 1
                    else:
                        year = base_year
                else:
                    year = base_year
                report_date = dt.date(year, month, day)
                reports.append(("result", summary, report_date))
            except ValueError:
                print(f"  警告: 無効な日付 {month}/{day}")
    
    # 日付なしのパターン（後方互換性のため）
    if not reports:
        # #日報計画（日付なし）
        for m in re.finditer(r"#日報計画\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
            summary = m.group(1).strip()[:50]
            if summary and not re.match(r"\d{1,2}/\d{1,2}", summary):
                reports.append(("plan", summary, None))
        
        # #日報結果（日付なし）
        for m in re.finditer(r"#日報結果\s*(?:要約[:：]\s*)?(.+?)(?:\n|$)", text):
            summary = m.group(1).strip()[:50]
            if summary and not re.match(r"\d{1,2}/\d{1,2}", summary):
                reports.append(("result", summary, None))
        
        # #日報（日付なし）
        for m in re.finditer(r"#日報\s+(.+?)(?:\n|$)", text):
            summary = m.group(1).strip()[:50]
            if (summary and "計画" not in summary and
                    "結果" not in summary and
                    not re.match(r"\d{1,2}/\d{1,2}", summary)):
                reports.append(("result", summary, None))
    
    return reports


# ===== Outlook 取得 =====
outlook = win32com.client.Dispatch("Outlook.Application")
namespace = outlook.GetNamespace("MAPI")

inbox = namespace.GetDefaultFolder(6)  # 6 = Inbox
target_folder = None
for f in inbox.Folders:
    if f.Name == OUTLOOK_FOLDER:
        target_folder = f
        break

if not target_folder:
    raise RuntimeError(f"Outlook フォルダが見つかりません: {OUTLOOK_FOLDER}")

processed = load_state()
new_processed = set(processed)

# ===== Excel 準備 =====
print(f"Excelファイル: {EXCEL_PATH}")
print(f"ファイル存在確認: {EXCEL_PATH.exists()}")
print()

if not EXCEL_PATH.exists():
    raise RuntimeError(f"Excelファイルが見つかりません: {EXCEL_PATH}")

try:
    wb = load_workbook(EXCEL_PATH, keep_vba=True, data_only=False)
except PermissionError:
    print("=" * 80)
    print("エラー: Excelファイルが開かれています")
    print("=" * 80)
    print()
    print("解決方法:")
    print("  1. Excelファイルを閉じてください")
    print("  2. もう一度このスクリプトを実行してください")
    print()
    print(f"ファイルパス: {EXCEL_PATH}")
    print("=" * 80)
    exit(1)
except Exception as e:
    print("=" * 80)
    print(f"エラー: Excelファイルの読み込みに失敗しました")
    print("=" * 80)
    print()
    print(f"詳細: {type(e).__name__}: {e}")
    print()
    print(f"ファイルパス: {EXCEL_PATH}")
    print("=" * 80)
    exit(1)

# 正しいシートを選択
target_sheet = None
for sheet_name in wb.sheetnames:
    if SHEET_NAME_PATTERN in sheet_name:
        target_sheet = sheet_name
        break

if target_sheet:
    ws = wb[target_sheet]
else:
    ws = wb.active
    print(f"警告: '{SHEET_NAME_PATTERN}'を含むシートが見つかりません。"
          f"'{ws.title}'を使用します。")


def find_row_by_date(ws, target_date):
    for r in range(1, ws.max_row + 1):
        v = ws[f"{DATE_COL}{r}"].value
        if isinstance(v, (dt.date, dt.datetime)):
            cell_date = v.date() if isinstance(v, dt.datetime) else v
            if cell_date == target_date:
                return r
    return None


# ===== メール処理 =====
print(f"メール処理開始 (フォルダ: {OUTLOOK_FOLDER})")
print(f"メール件数: {target_folder.Items.Count}")
print()

processed_count = 0
new_count = 0

for mail in target_folder.Items:
    entry_id = mail.EntryID
    
    if entry_id in processed:
        processed_count += 1
        continue

    # メールの受信日時を取得（デフォルト日付として使用）
    default_date = mail.ReceivedTime
    if isinstance(default_date, dt.datetime):
        default_date = default_date.date()
    else:
        print("✗ スキップ: 受信日時を取得できません")
        continue

    body = mail.Body or ""
    subject = mail.Subject or "(件名なし)"
    reports = extract_daily_reports(body, default_date)

    if reports:
        print(f"✓ 処理: {subject[:50]}")
        print(f"  抽出された項目: {len(reports)}件")
        
        # 日付ごとにグループ化して処理
        date_groups = {}
        for kind, summary, report_date in reports:
            # 日付が指定されていない場合はメール受信日を使用
            target_date = report_date if report_date else default_date
            
            if target_date not in date_groups:
                date_groups[target_date] = []
            date_groups[target_date].append((kind, summary))
        
        # 日付ごとにExcelに書き込み
        for target_date, items in date_groups.items():
            # メールの日付に対応する行を検索（日付表示行）
            date_row = find_row_by_date(ws, target_date)
            if date_row is None:
                print(f"  ✗ スキップ: {target_date} の日付行が見つかりません")
                continue
            
            # 実際の入力エリアは日付行の3行上から始まる
            input_start_row = date_row - 3
            
            print(f"  日付: {target_date} (日付行: {date_row}, "
                  f"入力開始行: {input_start_row})")
            
            # この日付ブロックの各列の次の空き行を追跡
            plan_row_offset = 0  # 計画列は0行目から
            result_row_offset = 0  # 結果列も0行目から
            
            for kind, summary in items:
                print(f"    - {kind}: {summary}")
                
                if kind == "plan":
                    # 計画列（C列）の空き行を探して書き込み
                    written = False
                    for offset in range(plan_row_offset, ROWS_PER_DAY):
                        target_row = input_start_row + offset
                        cell = ws[f"{PLAN_COL}{target_row}"]
                        if not cell.value:
                            cell.value = summary
                            print(f"      → セル {PLAN_COL}{target_row} に書き込み")
                            written = True
                            plan_row_offset = offset + 1  # 次は次の行から
                            break
                    
                    if not written:
                        # 空き行がない場合は最後の行に追記
                        target_row = input_start_row + ROWS_PER_DAY - 1
                        cell = ws[f"{PLAN_COL}{target_row}"]
                        if cell.value:
                            cell.value = f"{cell.value}\n{summary}"
                        else:
                            cell.value = summary
                        print(f"      → セル {PLAN_COL}{target_row} に追記"
                              f"（空き行なし）")
                        
                elif kind == "result":
                    # 結果列（F列）の空き行を探して書き込み
                    written = False
                    for offset in range(result_row_offset, ROWS_PER_DAY):
                        target_row = input_start_row + offset
                        cell = ws[f"{RESULT_COL}{target_row}"]
                        if not cell.value:
                            cell.value = summary
                            print(f"      → セル {RESULT_COL}{target_row} に書き込み")
                            written = True
                            result_row_offset = offset + 1  # 次は次の行から
                            break
                    
                    if not written:
                        # 空き行がない場合は最後の行に追記
                        target_row = input_start_row + ROWS_PER_DAY - 1
                        cell = ws[f"{RESULT_COL}{target_row}"]
                        if cell.value:
                            cell.value = f"{cell.value}\n{summary}"
                        else:
                            cell.value = summary
                        print(f"      → セル {RESULT_COL}{target_row} に追記"
                              f"（空き行なし）")

        new_processed.add(entry_id)
        new_count += 1
    else:
        print(f"✗ スキップ: {subject[:50]}")
        print("  理由: 日報タグまたは要約が見つかりません")

print()
print(f"処理結果: 新規 {new_count}件, スキップ済み {processed_count}件")
print()

# ===== 保存 =====
try:
    wb.save(EXCEL_PATH)
    save_state(new_processed)
    print("日報更新完了")
except PermissionError:
    print("=" * 80)
    print("エラー: Excelファイルの保存に失敗しました")
    print("=" * 80)
    print()
    print("原因: ファイルが他のプログラムで開かれています")
    print()
    print("解決方法:")
    print("  1. Excelファイルを閉じてください")
    print("  2. もう一度このスクリプトを実行してください")
    print()
    print(f"ファイルパス: {EXCEL_PATH}")
    print("=" * 80)
    print()
    print("注意: メールは処理されましたが、Excelへの書き込みは完了していません")
    exit(1)
except Exception as e:
    print("=" * 80)
    print(f"エラー: Excelファイルの保存に失敗しました")
    print("=" * 80)
    print()
    print(f"詳細: {type(e).__name__}: {e}")
    print()
    print(f"ファイルパス: {EXCEL_PATH}")
    print("=" * 80)
    exit(1)
