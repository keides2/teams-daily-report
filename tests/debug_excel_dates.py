"""
Excel日報ファイルの特定行の日付を確認するスクリプト
"""

import datetime as dt
from pathlib import Path
from openpyxl import load_workbook

# ⚠️ 注意: 実際の環境に合わせてパスを修正してください
EXCEL_PATH = Path("C:\\Users\\username\\株式会社サンプル\\部署名\\業務管理\\25年度個人別勤務台帳_サンプル.xlsx")
DATE_COL = "B"

print(f"Excelファイル: {EXCEL_PATH.name}")
print()

if not EXCEL_PATH.exists():
    print("エラー: Excelファイルが見つかりません")
    exit(1)

wb = load_workbook(EXCEL_PATH, keep_vba=True, data_only=True)
ws = wb.active

print(f"シート名: {ws.title}")
print()

today = dt.date.today()
print(f"今日の日付: {today} (type: {type(today)})")
print()

# 8行間隔でチェック (B24, B32, B40...)
target_rows = [24, 32, 40, 48, 56, 64, 72, 80, 88, 96]

print("行24周辺のセル内容を確認 (A-G列):")
print("-" * 80)
for r in range(22, 27):
    row_data = []
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        v = ws[f"{col}{r}"].value
        if v:
            row_data.append(f"{col}:{repr(v)}")
    if row_data:
        print(f"行{r}: {', '.join(row_data)}")

print()
print(f"{DATE_COL}列の日付を確認 (8行間隔):")
print("-" * 80)

found_today = False
found_row = None

for r in target_rows:
    v = ws[f"{DATE_COL}{r}"].value
    if v:
        print(f"行{r}: 値={repr(v)}, 型={type(v).__name__}", end="")
        
        if isinstance(v, (dt.date, dt.datetime)):
            date_obj = v.date() if isinstance(v, dt.datetime) else v
            print(f" → 日付={date_obj}", end="")
            
            if date_obj == today:
                print(" ★今日")
                found_today = True
                found_row = r
            else:
                print()
        else:
            print(" (日付型ではない)")
    else:
        print(f"行{r}: (空)")

print("-" * 80)

if found_today:
    print(f"\n✓ 今日の日付が行{found_row}に見つかりました")
else:
    print("\n✗ 今日の日付が見つかりませんでした")
    print("\n考えられる原因:")
    print("1. 今日の日付行がまだ追加されていない")
    print("2. 日付の形式が異なる (文字列など)")
    print("3. 検索範囲が不足している (もっと下の行にある)")
