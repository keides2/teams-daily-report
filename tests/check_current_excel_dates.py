"""
現在の.env設定でExcelの日付を確認
"""
import os
import datetime as dt
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook

# .env ファイルから環境変数を読み込み
load_dotenv()

# 現在の月に基づいてファイルパスを動的に生成
today = dt.date.today()
year = today.year
month = today.month

# 年度計算（4月始まり）
fiscal_year = year if month >= 4 else year - 1

# 年度内の月番号
if month >= 4:
    fiscal_month_num = month - 3
else:
    fiscal_month_num = month + 9

# 月フォルダ名
month_folder = f"{fiscal_month_num:02d}_{month}月"
month_name = f"{month}月"

# ===== Excelファイルパス設定 =====
excel_path_template = os.getenv("EXCEL_PATH", "")

if excel_path_template:
    excel_path_str = excel_path_template.format(
        fiscal_year=fiscal_year,
        month_folder=month_folder,
        month_name=month_name
    )
    EXCEL_PATH = Path(excel_path_str)
else:
    print("EXCEL_PATH環境変数が設定されていません")
    exit(1)

DATE_COL = os.getenv("DATE_COL", "B")
SHEET_NAME_PATTERN = f"{month}月"

print(f"Excelファイル: {EXCEL_PATH}")
print(f"ファイル存在確認: {EXCEL_PATH.exists()}")
print()

if not EXCEL_PATH.exists():
    print("エラー: Excelファイルが見つかりません")
    exit(1)

wb = load_workbook(EXCEL_PATH, data_only=True)

# 正しいシートを選択
target_sheet = None
for sheet_name in wb.sheetnames:
    if SHEET_NAME_PATTERN in sheet_name:
        target_sheet = sheet_name
        break

if target_sheet:
    ws = wb[target_sheet]
    print(f"使用シート: {target_sheet}")
else:
    ws = wb.active
    print(f"警告: '{SHEET_NAME_PATTERN}'を含むシートが見つかりません。'{ws.title}'を使用します。")

print()
print(f"日付列: {DATE_COL}")
print("=" * 80)

# 12/20から12/25あたりの日付を探す
target_dates = []
for day in range(20, 26):
    try:
        target_dates.append(dt.date(2025, 12, day))
    except ValueError:
        pass

print("対象日付範囲の検索:")
for target_date in target_dates:
    found = False
    for r in range(1, min(ws.max_row + 1, 300)):  # 最大300行まで
        v = ws[f"{DATE_COL}{r}"].value
        if isinstance(v, (dt.date, dt.datetime)):
            cell_date = v.date() if isinstance(v, dt.datetime) else v
            if cell_date == target_date:
                print(f"  {target_date} → 行 {r}")
                found = True
                break
    if not found:
        print(f"  {target_date} → 見つかりません")

print()
print("=" * 80)
print("200-220行目の日付データ:")
for r in range(200, 221):
    v = ws[f"{DATE_COL}{r}"].value
    if v:
        print(f"  行 {r}: {v} (型: {type(v).__name__})")
