"""
Excel日報ファイルの日付列を確認するスクリプト
"""

import datetime as dt
from pathlib import Path
from openpyxl import load_workbook

# ⚠️ 注意: 実際の環境に合わせてパスを修正してください
EXCEL_PATH = Path("C:\\Users\\username\\株式会社サンプル\\部署名\\業務管理\\25年度個人別勤務台帳_サンプル.xlsx")
DATE_COL = "B"

print(f"Excelファイル: {EXCEL_PATH}")
print(f"ファイル存在確認: {EXCEL_PATH.exists()}")
print()

if not EXCEL_PATH.exists():
    print("エラー: Excelファイルが見つかりません")
    exit(1)

wb = load_workbook(EXCEL_PATH, keep_vba=True, data_only=True)
ws = wb.active

print(f"シート名: {ws.title}")
print(f"最大行数: {ws.max_row}")
print()

today = dt.date.today()
print(f"今日の日付: {today}")
print()

print(f"{DATE_COL}列の日付を確認 (最初の30行):")
print("-" * 60)

found_today = False
for r in range(1, min(31, ws.max_row + 1)):
    v = ws[f"{DATE_COL}{r}"].value
    if v:
        date_str = str(v)
        if isinstance(v, (dt.date, dt.datetime)):
            date_obj = v.date() if isinstance(v, dt.datetime) else v
            date_str = f"{v} → {date_obj}"
            if date_obj == today:
                print(f"  行{r}: {date_str} ★今日")
                found_today = True
            else:
                print(f"  行{r}: {date_str}")
        else:
            print(f"  行{r}: {date_str} (日付型ではない)")

print("-" * 60)
if found_today:
    print("\n✓ 今日の日付が見つかりました")
else:
    print("\n✗ 今日の日付が見つかりませんでした")
    print("\n対応方法:")
    print("1. Excelファイルを開いて、今日の日付行を追加してください")
    print(f"2. または、スクリプトを修正して別の日付でテストしてください")
