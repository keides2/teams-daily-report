"""
テスト用サンプル Excel ファイルを生成するスクリプト。

Excel 構造（ROWS_PER_DAY=6, 各日付ブロック=8行）:
  input_start_row + 0  : 入力行1  (C列=計画, F列=実績)
  input_start_row + 1  : 入力行2
  input_start_row + 2  : 入力行3
  input_start_row + 3  : 日付行  (B列に日付)
  input_start_row + 4  : 入力行5
  input_start_row + 5  : 入力行6
  input_start_row + 6  : 余白
  input_start_row + 7  : 余白

生成ファイル: tests/sample_report.xlsx
シート名   : 2月  （SHEET_NAME_PATTERN と一致させる）
"""

import datetime as dt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent / "sample_report.xlsx"

# --- 構造定数（daily_report_writer.py と合わせる） ---
DATE_COL   = "B"
PLAN_COL   = "C"
RESULT_COL = "F"
ROWS_PER_DAY = 6
GAP_ROWS     = 2
BLOCK_SIZE   = ROWS_PER_DAY + GAP_ROWS   # = 8
DATE_OFFSET  = 3                          # 日付行 = input_start_row + 3

# --- 対象日付（2026年2月 平日のみサンプル） ---
TARGET_DATES = [
    dt.date(2026, 2, 19),  # 木
    dt.date(2026, 2, 20),  # 金
    dt.date(2026, 2, 21),  # 土（今日）
    dt.date(2026, 2, 23),  # 月（22は日曜）
    dt.date(2026, 2, 24),  # 火
]

HEADER_ROW = 1   # 1行目にヘッダー
DATA_START = 2   # データは2行目から


def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def make_sample_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "2月"

    # --- 列幅 ---
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 36

    # --- ヘッダー行 ---
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    headers = {
        "B": "日付",
        "C": "業務計画",
        "D": "区分",
        "E": "時間",
        "F": "業務実績",
    }
    for col, label in headers.items():
        cell = ws[f"{col}{HEADER_ROW}"]
        cell.value = label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- 日付ブロックを順に配置 ---
    date_fill  = PatternFill("solid", fgColor="D9E1F2")
    date_font  = Font(bold=True)
    input_fill = PatternFill("solid", fgColor="EBF3FB")

    for i, target_date in enumerate(TARGET_DATES):
        input_start_row = DATA_START + i * BLOCK_SIZE
        date_row        = input_start_row + DATE_OFFSET

        for offset in range(ROWS_PER_DAY):
            row = input_start_row + offset
            # 入力セル（C・F列）に薄い背景と枠線
            for col in [PLAN_COL, RESULT_COL]:
                cell = ws[f"{col}{row}"]
                cell.fill = input_fill
                cell.border = thin_border()
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 日付セル（B列）
        date_cell = ws[f"{DATE_COL}{date_row}"]
        date_cell.value = dt.datetime(
            target_date.year, target_date.month, target_date.day
        )
        date_cell.number_format = "MM/DD(AAA)"
        date_cell.fill = date_fill
        date_cell.font = date_font
        date_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 行の高さ
        for offset in range(ROWS_PER_DAY):
            ws.row_dimensions[input_start_row + offset].height = 18
        # 余白行
        for gap in range(GAP_ROWS):
            ws.row_dimensions[input_start_row + ROWS_PER_DAY + gap].height = 6

    wb.save(OUTPUT_PATH)
    print(f"サンプル Excel を生成しました: {OUTPUT_PATH}")
    print(f"シート名: {ws.title}")
    print()
    print("日付行の確認:")
    for i, d in enumerate(TARGET_DATES):
        input_start = DATA_START + i * BLOCK_SIZE
        date_row    = input_start + DATE_OFFSET
        print(f"  {d}  → B{date_row}  (入力開始: 行{input_start}〜{input_start + ROWS_PER_DAY - 1})")


if __name__ == "__main__":
    make_sample_excel()
