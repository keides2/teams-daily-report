"""
Excelファイルから日付を読み込むテスト
"""

import datetime as dt
from pathlib import Path
from openpyxl import load_workbook

# ⚠️ 注意: 実際の環境に合わせてパスを修正してください
EXCEL_PATH = Path("C:\\Users\\username\\OneDrive - 株式会社サンプル\\部署名-01.庶務事項\\業務管理\\業務内容報告書\\2025年度\\09_12月\\チームA\\（12月_山田太郎）業務内容報告書.xlsm")
DATE_COL = "B"

print("★ 重要: Excelファイルを閉じてから実行してください ★\n")

print("data_only=True で読み込み:")
print("-" * 80)

# read_onlyモードで開く
wb = load_workbook(EXCEL_PATH, read_only=False, keep_vba=False, data_only=False)

print(f"全シート一覧: {wb.sheetnames}")
print()

# 「12月」を含むシートを探す
target_sheet = None
for sheet_name in wb.sheetnames:
    if "12月" in sheet_name or "2025年12月" in sheet_name:
        target_sheet = sheet_name
        break

if target_sheet:
    ws = wb[target_sheet]
    print(f"★ '{ws.title}' シートを使用します")
else:
    ws = wb.active
    print(f"★ アクティブシート '{ws.title}' を使用します (12月シートが見つかりません)")

print()
print(f"シート: {ws.title}")
print()

today = dt.date.today()
print(f"今日の日付: {today}\n")

# B24, B32, B40 を確認
test_rows = [24, 32, 40, 48]

for r in test_rows:
    cell = ws[f"{DATE_COL}{r}"]
    v = cell.value
    print(f"行{r}:")
    print(f"  値: {repr(v)}, 型: {type(v).__name__}")
    print(f"  数式: {cell.data_type}, number_format: {cell.number_format}")
    
    if isinstance(v, (dt.date, dt.datetime)):
        cell_date = v.date() if isinstance(v, dt.datetime) else v
        print(f"  → 日付として解釈: {cell_date}")
        if cell_date == today:
            print("  ★ 今日です!")
    elif isinstance(v, (int, float)):
        # Excelのシリアル値として解釈してみる
        try:
            # Excelの日付シリアル値 (1900年1月1日からの日数)
            excel_date = dt.datetime(1899, 12, 30) + dt.timedelta(days=v)
            print(f"  → Excelシリアル値として解釈: {excel_date.date()}")
            if excel_date.date() == today:
                print("  ★ 今日です!")
        except:
            pass
    print()

# より広い列範囲を確認
print("\n行24の全列 (A-K) を確認:")
print("-" * 80)
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    cell = ws[f"{col}24"]
    v = cell.value
    if v is not None:
        formula = f", 数式:{cell.value if hasattr(cell, 'value') else 'N/A'}"
        print(f"{col}24: {repr(v)} (型:{type(v).__name__}, format:{cell.number_format}){formula}")
        if isinstance(v, (dt.date, dt.datetime)):
            cell_date = v.date() if isinstance(v, dt.datetime) else v
            print(f"    → {cell_date}")

print("\n★ Excelファイルを開いて「Ctrl+S」で保存してから再実行してください ★")

# 全行スキャンして12/24を探す
print("\n12/24を持つ行を全体から検索:")
print("-" * 80)

found = False
for r in range(1, min(ws.max_row + 1, 300)):
    v = ws[f"{DATE_COL}{r}"].value
    if isinstance(v, (dt.date, dt.datetime)):
        cell_date = v.date() if isinstance(v, dt.datetime) else v
        if cell_date.month == 12 and cell_date.day == 24:
            print(f"✓ 見つかりました! 行{r}: {cell_date}")
            found = True
            
            # その行の内容も表示
            print(f"\n  行{r}の内容:")
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                cell_val = ws[f"{col}{r}"].value
                if cell_val:
                    val_str = str(cell_val)[:40]
                    print(f"    {col}: {val_str}")
            break

if not found:
    print("✗ 12/24の行が見つかりませんでした")
    print(f"  検索範囲: 行1〜{min(ws.max_row, 300)}")
    print(f"  最大行数: {ws.max_row}")
