"""
ExcelファイルのB列の日付を2025年から2026年に変更するスクリプト
複数の月のファイルを一括処理
"""
import os
import datetime as dt
from pathlib import Path
from openpyxl import load_workbook
from dotenv import load_dotenv

# 環境変数の読み込み
load_dotenv()

# 処理対象の月リスト（1月、2月、3月）
TARGET_MONTHS = [1, 2, 3]

# 現在の年度を取得
today = dt.date.today()
year = today.year
fiscal_year = year if today.month >= 4 else year - 1

# Excelファイルパステンプレート
excel_path_template = os.getenv("EXCEL_PATH", "")

if not excel_path_template:
    print("エラー: EXCEL_PATH 環境変数が設定されていません")
    exit(1)

# 各月のExcelファイルパスを生成
excel_paths = []
for month in TARGET_MONTHS:
    # 年度内の月番号
    if month >= 4:
        fiscal_month_num = month - 3
    else:
        fiscal_month_num = month + 9
    
    # 月フォルダ名
    month_folder = f"{fiscal_month_num:02d}_{month}月"
    month_name = f"{month}月"
    
    # パス生成
    excel_path_str = excel_path_template.format(
        fiscal_year=fiscal_year,
        month_folder=month_folder,
        month_name=month_name
    )
    EXCEL_PATH = Path(excel_path_str)
    excel_paths.append((month, EXCEL_PATH))
    excel_paths.append((month, EXCEL_PATH))

# 各ファイルを処理
total_files = len(excel_paths)
processed_files = 0

print("=" * 80)
print("Excel年度修正ツール: 2025年 → 2026年")
print("=" * 80)
print(f"\n処理対象: {total_files}ファイル")
print()

for month, EXCEL_PATH in excel_paths:
    print(f"\n{'=' * 80}")
    print(f"【{month}月のファイル】")
    print(f"パス: {EXCEL_PATH}")
    
    if not EXCEL_PATH.exists():
        print("⚠️  ファイルが見つかりません - スキップ")
        continue
    
    try:
        # Excelファイルを開く
        print("読み込み中...")
        wb = load_workbook(EXCEL_PATH, keep_vba=True, data_only=False)
        
        # シート名のパターンに一致するシートを探す
        SHEET_NAME_PATTERN = f"{month}月"
        ws = None
        for sheet_name in wb.sheetnames:
            if SHEET_NAME_PATTERN in sheet_name:
                ws = wb[sheet_name]
                print(f"使用シート: {sheet_name}")
                break
        
        if ws is None:
            print(f"⚠️  {SHEET_NAME_PATTERN}を含むシートが見つかりません - スキップ")
            wb.close()
            continue
        
        # B列の日付を変更
        DATE_COL = os.getenv("DATE_COL", "B")
        updated_count = 0
        
        print("日付を変更中...")
        
        for row in range(1, ws.max_row + 1):
            cell = ws[f"{DATE_COL}{row}"]
            
            if isinstance(cell.value, dt.datetime) or isinstance(cell.value, dt.date):
                date_value = cell.value
                
                # 2025年の日付を2026年に変更
                if date_value.year == 2025:
                    new_date = date_value.replace(year=2026)
                    cell.value = new_date
                    if updated_count < 5:  # 最初の5件のみ表示
                        print(f"  行 {row}: {date_value.strftime('%Y-%m-%d')} → {new_date.strftime('%Y-%m-%d')}")
                    updated_count += 1
        
        if updated_count > 5:
            print(f"  ... 他 {updated_count - 5}件")
        
        print(f"変更件数: {updated_count}件")
        
        # 保存
        if updated_count > 0:
            print("保存中...")
            wb.save(EXCEL_PATH)
            print("✅ 保存完了")
            processed_files += 1
        else:
            print("ℹ️  変更なし")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ エラー: {type(e).__name__}: {e}")
        continue

print(f"\n{'=' * 80}")
print(f"処理完了: {processed_files}/{total_files}ファイル")
print("=" * 80)
